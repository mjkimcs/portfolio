#urlopen함수보다 더 안정적인 requests.get함수 사용(code.text 주의)
#이미지 크롤링 심화: .jpg가 아닌 .git가 있을 경우

from bs4 import BeautifulSoup #pip install beautifulsoup4
import requests #pip install requests
import os
import openpyxl #pip install openpyxl
import urllib.parse as par
from openpyxl.drawing.image import Image

if not os.path.exists("./쿠팡상품정보.xlsx"):
  o = openpyxl.Workbook()
  o.save("./쿠팡상품정보.xlsx")

o = openpyxl.load_workbook("./쿠팡상품정보.xlsx")
sheet = o.active
sheet.cell(row=1, column=1).value = "페이지"
sheet.cell(row=1, column=2).value = "이미지"
sheet.cell(row=1, column=3).value = "상품명"
sheet.cell(row=1, column=4).value = "평점"
sheet.cell(row=1, column=5).value = "리뷰 수"
sheet.cell(row=1, column=6).value = "링크"
sheet.column_dimensions["A"].width = "8"
sheet.column_dimensions["B"].width = "31"
sheet.column_dimensions["C"].width = "60"
row_num = 2

search = input("검색>>")
encoded = par.quote(search)

if not os.path.exists("./쿠팡상품이미지"):
    os.mkdir("./쿠팡상품이미지")

pg_num = 1
while True:
    if pg_num == 6:
        break
    print("{}페이지".format(pg_num))
    sheet.cell(row=row_num, column=1).value = "{}페이지".format(pg_num)
    url = "https://www.coupang.com/np/search?q={}&channel=user&component=&eventCategory=SRP&trcid=&traid=&sorter=saleCountDesc&minPrice=&maxPrice=&priceRange=&filterType=&listSize=36&filter=&isPriceRange=false&brand=&offerCondition=&rating=0&page={}&rocketAll=false&searchIndexingToken=&backgroundColor=".format(encoded,pg_num)
    code = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}) #urlopen함수와 동일
    soup = BeautifulSoup(code.text, "html.parser") #code.text
    all = soup.select("li.search-product")
    for i in all:
        ad = i.select("span.ad-badge")
        img = i.select_one("img.search-product-wrap-img")
        name = i.select_one("div.name")
        rating = i.select_one("em.rating")
        review = i.select_one("span.rating-total-count")
        link = i.select_one("a")
        if len(ad) != 0: #광고상품 제거
            continue
        if "신지모루" in name.string:
            continue
        if rating:
            rating.string
        else:
            continue
        if review:
            review.string.replace("(", "").replace(")", "")
        else:
            continue

        if float(rating.string) >= 4.5 and float(review.string.replace("(", "").replace(")", "")) >= 500:
            img_f = "https:" + img.attrs["src"]
            if ".gif" in img_f:  #파일명에 "gif"가 들어가 있는 것들이 있음
                img_f = "https:" + img.attrs["data-img-src"]
            with open("./쿠팡상품이미지/{}.jpg".format(int(row_num) - 1), "wb") as w:
                w.write(requests.get(img_f).content)
            name_f = name.string
            rating_f = rating.string
            review_f = review.string.replace("(", "").replace(")", "")
            link_f = "https://www.coupang.com" + link["href"]
            print(name_f, rating_f, review_f, link_f)
            img_for_excel = Image("./쿠팡상품이미지/{}.jpg".format(int(row_num)-1))
            sheet.add_image(img_for_excel, "B{}".format(row_num))
            sheet.row_dimensions[row_num].height = "190"
            sheet.cell(row=row_num, column=3).value = name_f
            sheet.cell(row=row_num, column=4).value = rating_f
            sheet.cell(row=row_num, column=5).value = review_f
            sheet.cell(row=row_num, column=6).value = link_f
            o.save("./쿠팡상품정보.xlsx")
            row_num += 1
    pg_num += 1