from bs4 import BeautifulSoup
import requests
import os
import openpyxl
import urllib.request as req
import urllib.parse as par
import json
import time
import re
import openpyxl
import os

if not os.path.exists("./petition.xlsx"):
    o = openpyxl.Workbook()
    o.save("./petition.xlsx")

o = openpyxl.load_workbook("./petition.xlsx")
sheet = o.active

sheet.cell(row=1, column=1).value = "title"
sheet.cell(row=1, column=2).value = "vote"
sheet.cell(row=1, column=3).value = "category"
sheet.cell(row=1, column=4).value = "startdate"
sheet.cell(row=1, column=5).value = "enddate"
sheet.cell(row=1, column=6).value = "content"

pg_num = 1
while True:
    if pg_num == 2:
        break

    url = "https://www1.president.go.kr/api/petitions/list"
    post_data = {
    "c": "0",
    "only": "2",
    "page": str(pg_num),  # 문자열 자료형만 들어가야 하기 때문에 str() 사용
    "order": "1"
    }
    # Request Headers 복붙 시 주의할 점은 ":authority", ":method", ":path"처럼 앞에 콜론(:)이 붙어있는 것들은 제외
    post_header = {
    "accept": "application/json, text/javascript, */*; q=0.01",
    "accept-encoding": "gzip, deflate, br",
    "accept-language": "ko-KR,ko;q=0.9,en-US;q=0.8,en;q=0.7",
    "content-length": "10",
    "content-type": "application/x-www-form-urlencoded; charset=UTF-8",
    "cookie": "_gid=GA1.3.1411726365.1613402116; _ga=GA1.3.1951266142.1613300336; _ga_1PBY96CP0Y=GS1.1.1613402115.2.1.1613402180.0",
    "origin": "https://www1.president.go.kr",
    "referer": "https://www1.president.go.kr/petitions/?c=0&only=2&page=3&order=1",
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-origin",
    "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_16_0) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.106 Whale/2.8.108.15 Safari/537.36",
    "x-requested-with": "XMLHttpRequest"
    }

    result = requests.post(url, data=post_data, headers=post_header)  # urlopen함수와 동일
    result = json.loads(result.text)  # json->딕셔너리 자료형

    row_num = 2
    for i in result["item"]:
        url = "https://www1.president.go.kr/petitions/{}".format(i["id"])
        code = requests.get(url, headers={"User-Agent": "Mozilla/5.0"})  # urlopen함수와 동일
        soup = BeautifulSoup(code.text, "html.parser")  # code.text
        title = soup.select_one("h3.petitionsView_title")
        vote = soup.select_one("span.counter")
        box = soup.select("ul.petitionsView_info_list > li")
        content = soup.select_one("div.View_write")
        content_re = re.sub( "(\r)|(\n)|(\t)", " ", content.text.strip() )
        print(title.string)
        sheet.cell(row=row_num, column=1).value = title.string
        print(vote.string)
        sheet.cell(row=row_num, column=2).value = vote.string
        print(box[0].text[4:])
        sheet.cell(row=row_num, column=3).value = box[0].text[4:]
        print(box[1].text[4:])
        sheet.cell(row=row_num, column=4).value = box[1].text[4:]
        print(box[2].text[4:])
        sheet.cell(row=row_num, column=5).value = box[2].text[4:]
        print(content_re)
        sheet.cell(row=row_num, column=6).value = content_re
        print("-----------------")
        o.save("./petition.xlsx")
        row_num += 1

    pg_num += 1
    time.sleep(2)  # 청와대 너무 빠르게 크롤링할 시 서버에서 차단


