import urllib.request as req
from bs4 import BeautifulSoup
import urllib.parse as par
import pandas as pd
import openpyxl
import os
import matplotlib.pyplot as plt

if not os.path.exists("./로또.xlsx"):
    o = openpyxl.Workbook()
    o.save("./로또.xlsx")

o = openpyxl.load_workbook("./로또.xlsx")
sheet = o.active

sheet.cell(row=1, column=1).value = "로또번호"

row_num = 2
lotto_turn = 0
while True:
    lotto_turn += 1
    query = par.quote("로또" + str(lotto_turn) + "회")
    url = "https://search.naver.com/search.naver?sm=top_hty&fbm=0&ie=utf8&query={}".format(query)
    code = req.urlopen(url)
    soup = BeautifulSoup(code, "html.parser")
    nums = soup.select("div.num_box > span")

    if len(nums) == 0:
        break

    print("{}회차 :".format(lotto_turn), end=" ")
    bonus = nums[7]
    for i in nums[:6]:
        print(i.string, end=" ")
        sheet.cell(row=row_num, column=1).value = i.string
        row_num += 1
    print(bonus.string, end=" ")
    print()
    sheet.cell(row=row_num, column=1).value = bonus.string
    o.save("./로또.xlsx")
    row_num += 1

df = pd.read_excel("./로또.xlsx")
data = df["로또번호"].value_counts()
df2 = pd.DataFrame([data], index=["빈도 수"])
df2 = df2.transpose()
model = df2.plot(kind='bar', figsize=(15, 10))
model.get_legend().remove()
plt.show()