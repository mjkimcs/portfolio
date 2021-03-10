from selenium import webdriver
import time
import random
import openpyxl
import os
import unicodedata #한글 자모음 분리현상 해결

if not os.path.exists("./인스타그램.xlsx"):
    book = openpyxl.Workbook()
    book.save("./인스타그램.xlsx")

book = openpyxl.load_workbook("./인스타그램.xlsx")
# sheet = book["Sheet1"]
sheet = book.active #Sheet 지정없이 그냥 자동으로 열리는 현재 Sheet

hash_tag = input("해시태그 입력>>")

# option = webdriver.ChromeOptions()
# option.add_argument("headless")
b = webdriver.Chrome("./chromedriver") #, options=option
b.get("https://www.instagram.com/accounts/login/?source=auth_switcher")
time.sleep(3)

id = b.find_element_by_name("username")
id.send_keys("2min_kim")
pw = b.find_element_by_name("password")
pw.send_keys("1234")
b.find_element_by_css_selector("div.Igw0E.IwRSH.eGOV_._4EzTm.bkEs3.CovQj.jKUp7.DhRcB").click()
time.sleep(4)

url = "https://www.instagram.com/explore/tags/{}/".format(hash_tag)
b.get(url)
time.sleep(7)

b.find_element_by_css_selector("div._9AhH0").click()
time.sleep(3)

row_num = 1
while True:
    value = like.get_attribute("aria-label")
    next = b.find_element_by_css_selector("a._65Bje.coreSpriteRightPaginationArrow")

    content = b.find_element_by_css_selector("div.C4VMK > span")
    content_norm = unicodedata.normalize("NFC", content.text) #한글 자모음 분리현상 해결

    sheet.cell(row=row_num, column=1).value = nick_name.text
    sheet.cell(row=row_num, column=2).value = content_norm
    row_num += 1
    book.save("./인스타그램.xlsx")

    if value == "좋아요": #좋아요가 안 눌려져 있다면
        like.click()
        time.sleep(random.randint(2,5) + random.random()) #인스타그램을 속이기 위한 랜덤한 소수
        next.click()
        time.sleep(random.randint(2,5) + random.random())
    elif value == "좋아요 취소": #좋아요가 눌려져 있다면
        next.click()
        time.sleep(random.randint(2,5) + random.random())